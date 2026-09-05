#!/usr/bin/env python3
"""Split long BOM Designator cells into JLCPCB-compatible continuation rows."""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("Missing dependency: install it with 'python3 -m pip install openpyxl'.", file=sys.stderr)
    raise SystemExit(1)


DEFAULT_STEM = "BOM_Analog_Matrix_Inverted_MAML_IMC_2026-09-05"
MAX_DESIGNATOR_CHARS = 2047


def split_designators(value: str, maximum: int) -> list[str]:
    """Return comma-delimited groups no longer than maximum characters."""
    designators = [designator.strip() for designator in value.split(",")]
    if not designators or any(not designator for designator in designators):
        raise ValueError("Designator contains an empty entry between commas.")

    sections: list[str] = []
    section = ""
    for designator in designators:
        if len(designator) > maximum:
            raise ValueError(
                f"A single designator is {len(designator)} characters, exceeding {maximum}."
            )
        candidate = designator if not section else f"{section},{designator}"
        if len(candidate) > maximum:
            sections.append(section)
            section = designator
        else:
            section = candidate
    sections.append(section)
    return sections


def process_rows(rows: Iterable[Iterable[object]], maximum: int) -> list[list[object]]:
    rows = [list(row) for row in rows]
    if not rows:
        raise ValueError("The BOM is empty.")

    header = rows[0]
    try:
        designator_index = header.index("Designator")
        quantity_index = header.index("Quantity")
    except ValueError as error:
        raise ValueError("The BOM must contain 'Designator' and 'Quantity' columns.") from error

    output = [header]
    for source_row, row in enumerate(rows[1:], start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        if len(row) <= designator_index:
            raise ValueError(f"Row {source_row} has no Designator cell.")
        designator = "" if row[designator_index] is None else str(row[designator_index]).strip()
        if not designator:
            output.append(row)
            continue

        sections = split_designators(designator, maximum)
        for section in sections:
            continuation = row.copy()
            continuation[designator_index] = section
            continuation[quantity_index] = len(section.split(","))
            output.append(continuation)
    return output


def read_csv(path: Path) -> tuple[list[list[str]], str, str]:
    for encoding in ("utf-16", "utf-8-sig", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as bom_file:
                header = bom_file.readline()
                bom_file.seek(0)
                delimiter = "\t" if header.count("\t") >= header.count(",") else ","
                return list(csv.reader(bom_file, delimiter=delimiter)), encoding, delimiter
        except UnicodeError:
            continue
    raise ValueError(f"Could not decode {path.name} as UTF-16 or UTF-8.")


def write_csv(path: Path, rows: list[list[object]], encoding: str, delimiter: str) -> None:
    with path.open("w", encoding=encoding, newline="") as bom_file:
        writer = csv.writer(bom_file, delimiter=delimiter, lineterminator="\n")
        writer.writerows(rows)


def read_xlsx(path: Path) -> list[list[object]]:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def write_xlsx(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_Designator_max_2047_char{input_path.suffix}")


def process_csv(path: Path, maximum: int) -> Path:
    rows, encoding, delimiter = read_csv(path)
    destination = output_path(path)
    write_csv(destination, process_rows(rows, maximum), encoding, delimiter)
    return destination


def process_xlsx(path: Path, maximum: int) -> Path:
    destination = output_path(path)
    write_xlsx(destination, process_rows(read_xlsx(path), maximum))
    return destination


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", type=Path, default=Path(f"{DEFAULT_STEM}.csv"))
    parser.add_argument("--xlsx", type=Path, default=Path(f"{DEFAULT_STEM}.xlsx"))
    parser.add_argument("--max-chars", type=int, default=MAX_DESIGNATOR_CHARS)
    arguments = parser.parse_args()

    if arguments.max_chars < 1:
        parser.error("--max-chars must be positive")

    for processor, source in ((process_csv, arguments.csv), (process_xlsx, arguments.xlsx)):
        if not source.is_file():
            print(f"Skipping missing input: {source}", file=sys.stderr)
            continue
        destination = processor(source, arguments.max_chars)
        print(f"Created {destination}")


if __name__ == "__main__":
    main()