#!/usr/bin/env python3
"""Verify that a sectioned BOM preserves every original component reference."""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: install it with 'python3 -m pip install openpyxl'.", file=sys.stderr)
    raise SystemExit(1)


DEFAULT_STEM = "BOM_Analog_Matrix_Inverted_MAML_IMC_2026-09-05"
MAX_DESIGNATOR_CHARS = 2047


def read_csv(path: Path) -> list[list[object]]:
    for encoding in ("utf-16", "utf-8-sig", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as bom_file:
                header = bom_file.readline()
                bom_file.seek(0)
                delimiter = "\t" if header.count("\t") >= header.count(",") else ","
                return list(csv.reader(bom_file, delimiter=delimiter))
        except UnicodeError:
            continue
    raise ValueError(f"Could not decode {path.name} as UTF-16 or UTF-8.")


def read_xlsx(path: Path) -> list[list[object]]:
    sheet = load_workbook(path, data_only=False).active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def value(cell: object) -> str:
    return "" if cell is None else str(cell).strip()


def nonblank_rows(rows: Iterable[Iterable[object]]) -> list[list[object]]:
    return [list(row) for row in rows if any(cell not in (None, "") for cell in row)]


def component_key(row: list[object], excluded: set[int]) -> tuple[str, ...]:
    return tuple(value(cell) for index, cell in enumerate(row) if index not in excluded)


def references(designator: str) -> list[str]:
    return [reference.strip() for reference in designator.split(",") if reference.strip()]


def verify(original: list[list[object]], sectioned: list[list[object]], maximum: int) -> list[str]:
    errors: list[str] = []
    original = nonblank_rows(original)
    sectioned = nonblank_rows(sectioned)
    if not original or not sectioned:
        return ["One or both BOM files are empty."]

    original_header, sectioned_header = original[0], sectioned[0]
    if original_header != sectioned_header:
        return ["Headers differ between the original and generated BOM files."]
    try:
        designator_index = original_header.index("Designator")
        quantity_index = original_header.index("Quantity")
    except ValueError:
        return ["Both BOM files must contain 'Designator' and 'Quantity' columns."]

    excluded = {designator_index, quantity_index}
    original_references: Counter[tuple[tuple[str, ...], str]] = Counter()
    sectioned_references: Counter[tuple[tuple[str, ...], str]] = Counter()

    for row_number, row in enumerate(original[1:], start=2):
        if len(row) <= designator_index:
            errors.append(f"Original row {row_number} has no Designator cell.")
            continue
        key = component_key(row, excluded)
        for reference in references(value(row[designator_index])):
            original_references[(key, reference)] += 1

    for row_number, row in enumerate(sectioned[1:], start=2):
        if len(row) <= max(designator_index, quantity_index):
            errors.append(f"Generated row {row_number} has too few columns.")
            continue
        designator = value(row[designator_index])
        row_references = references(designator)
        if not designator or len(row_references) != len(designator.split(",")):
            errors.append(f"Generated row {row_number} has an empty Designator entry.")
            continue
        if len(designator) > maximum:
            errors.append(
                f"Generated row {row_number} has a {len(designator)}-character Designator cell."
            )
        try:
            quantity = int(value(row[quantity_index]))
        except ValueError:
            errors.append(f"Generated row {row_number} has a non-integer Quantity.")
        else:
            if quantity != len(row_references):
                errors.append(
                    f"Generated row {row_number} has Quantity {quantity}, "
                    f"but {len(row_references)} Designators."
                )
        key = component_key(row, excluded)
        for reference in row_references:
            sectioned_references[(key, reference)] += 1

    missing = original_references - sectioned_references
    extra = sectioned_references - original_references
    if missing:
        errors.append(f"Missing {sum(missing.values())} original Designator occurrence(s).")
    if extra:
        errors.append(f"Found {sum(extra.values())} extra or mismatched Designator occurrence(s).")
    return errors


def default_sectioned_path(path: Path) -> Path:
    return path.with_name(f"{path.stem}_Designator_max_2047_char{path.suffix}")


def verify_file_pair(label: str, reader, original_path: Path, sectioned_path: Path, maximum: int) -> bool:
    if not original_path.is_file() or not sectioned_path.is_file():
        print(f"{label}: SKIPPED (missing {original_path if not original_path.is_file() else sectioned_path})")
        return False
    errors = verify(reader(original_path), reader(sectioned_path), maximum)
    if errors:
        print(f"{label}: FAILED")
        for error in errors:
            print(f"  - {error}")
        return False
    print(f"{label}: OK")
    return True


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", type=Path, default=Path(f"{DEFAULT_STEM}.csv"))
    parser.add_argument("--csv-sectioned", type=Path)
    parser.add_argument("--xlsx", type=Path, default=Path(f"{DEFAULT_STEM}.xlsx"))
    parser.add_argument("--xlsx-sectioned", type=Path)
    parser.add_argument("--max-chars", type=int, default=MAX_DESIGNATOR_CHARS)
    arguments = parser.parse_args()
    if arguments.max_chars < 1:
        parser.error("--max-chars must be positive")

    results = [
        verify_file_pair(
            "CSV", read_csv, arguments.csv,
            arguments.csv_sectioned or default_sectioned_path(arguments.csv), arguments.max_chars,
        ),
        verify_file_pair(
            "XLSX", read_xlsx, arguments.xlsx,
            arguments.xlsx_sectioned or default_sectioned_path(arguments.xlsx), arguments.max_chars,
        ),
    ]
    if not all(results):
        raise SystemExit(1)


if __name__ == "__main__":
    main()